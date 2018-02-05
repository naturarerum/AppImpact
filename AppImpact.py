"""
   AppImpact v 1.0
   auteur : Olivier Lopez
   Date : 02-2018.

"""

import datetime
import tkinter as tk
from tkinter import HORIZONTAL
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from tkinter.ttk import Progressbar
from openpyxl import Workbook

import openpyxl


class MainGui:
    def __init__(self):

        pass

    def create_widgets(self):
        # creer un Frame principal
        self.release = tk.StringVar()
        main_frame = ttk.LabelFrame(self.root)
        main_frame.grid(column=0, row=0)

        # Frame boutons
        btn_Frame = ttk.LabelFrame(self.root)
        btn_Frame.grid(column=2, row=0)

        # Création du label Saisie de la release
        lbl_release = ttk.Label(main_frame, text="Release             : ").grid(column=0, row=0)

        # Création du label Fichier input
        lbl_ficinput = ttk.Label(main_frame, text="Input file          : ").grid(column=0, row=1)

        # Création du label Fichier output
        lbl_ficoutput = ttk.Label(main_frame, text="Output file       : ").grid(column=0, row=2)

        # Création du widget saisie de la release
        saisie_release = ttk.Entry(main_frame, width=15, textvariable=self.release)
        saisie_release.grid(column=2, row=0)
        # TODO : Remplacer par une liste déroulante auto-populée

        # Création d'un widget bouton pour lancer le traitement
        btn_start = ttk.Button(btn_Frame, text="Start", command=self.traitement).grid(column=2, row=0)

        # Création d'un widget bouton pour selectionner le fichier input
        btn_selectin = ttk.Button(btn_Frame, text="Select", command=self.openfic).grid(column=2, row=1)

        # Création d'un widget bouton pour selectionner le fichier output
        btn_selectout = ttk.Button(btn_Frame, text="Select", command=self.savefic).grid(column=2, row=2)

        # Création progressbar

    def openfic(self):
        try:
            self.ficin = filedialog.askopenfilename()
        except IOError:
            messagebox.showerror("Erreur", "le fichier est invalide")
        return self.ficin

    def savefic(self):
        try:
            self.ficout= filedialog.asksaveasfilename()
        except IOError:
            messagebox.showerror("Erreur", "le fichier est invalide")
        return self.ficout

    def traitement(self):
        # TODO : Ajouter une exception si release pas saisie
        
         liste_atraiter = self.traite_impact(self.ficin)
         liste_output = self.extrait_app(liste_atraiter)
         self.create_wb(liste_output)

    def traite_impact(self, filename):
        fichier = filename
        rel = self.release.get()  # récupere la saisie de l'utilisateur
        type_s = ['New', 'Update', 'Reuse']
        try:
            wb = openpyxl.load_workbook(fichier, data_only=True)
        except openpyxl.utils.exceptions.InvalidFileException:
            messagebox.showerror("Erreur", "le fichier doit etre au format xlsx")
        sheet = wb.get_sheet_by_name("Statut des services")
        maxrow = sheet.max_row
        row_list = []
        final_list = []
        for row_index in range(2, maxrow):
            flag_type = False
            flag_mois = False
            flag_application = False
            projet = ''
            type_service = ''
            nom_service = ''
            mois_release = ''
            application = ''
            # print('Row: ' + str(row_index))
            projet = sheet.cell(row=row_index, column=1).value
            type_service = sheet.cell(row=row_index, column=2).value
            if type_service in type_s:
                flag_type = True
            nom_service = sheet.cell(row=row_index, column=3).value
            mois_release = sheet.cell(row=row_index, column=10).value
            if mois_release == rel:
                flag_mois = True
            application = sheet.cell(row=row_index, column=25).value
            if application is not None:
                flag_application = True
            if flag_mois and flag_type and flag_application:
                row_list = [projet, type_service, nom_service, mois_release, application]
                print(row_list)
                final_list.append(row_list)
        return final_list

    def extrait_app(self, liste_atraiter):

        res2= []
        i = 0
        for i in range(len(liste_atraiter)):
            j = 0
            print('i = :  ', i, liste_atraiter[i])
            elem = str((liste_atraiter[i][4]))
            chaine = elem.split(",")
            for j in range(len(chaine)):
                res = []
                res.append(liste_atraiter[i][0])
                res.append(liste_atraiter[i][1])
                res.append(liste_atraiter[i][2])
                res.append(liste_atraiter[i][3])
                res.append(chaine[j])
                print("res", res)
                res2.append(res)
        print("res2", res2)
        print('----------------------------------------------------------------------------------------------')
        return res2

    def create_wb(self, liste_output):
        i = 0
        j = 1
        k = 0
        wb_res = Workbook()
        ws1 = wb_res.active
        ws1.title = str(datetime.date.today())
        dest_filename = self.ficout
        header = ['Projet', 'Type service', 'Service', 'Release', 'App. Impactée']
        for i in range(1, len(header) + 1):
            ws1.cell(row=1, column=i).value = header[i - 1]
        for j, line in enumerate(liste_output):
           for k, line in enumerate(line):
                ws1.cell(row=j + 2, column=k+1).value = line
        wb_res.save(dest_filename)

    def app(self):
        self.root = tk.Tk()
        self.root.title("AppImpact")
        # self.root.iconbitmap(r'C:\Python34\DLLs\pyc.ico')
        self.root.geometry("280x125")
        self.create_widgets()
        # self.progress = Progressbar(self, orient=HORIZONTAL, length=100, mode='indeterminate')
        self.root.mainloop()

# Main
if __name__ == '__main__':
    mg = MainGui()
    mg.app()


